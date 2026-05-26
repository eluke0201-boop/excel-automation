"""
EC事業者向け 受注データ統合・集計ツール

複数プラットフォーム（楽天市場・Amazon・自社BASE等）から出力された
受注CSVを自動で統合し、商品別・プラットフォーム別の集計レポートを
グラフ付きのExcelで出力する。

使い方:
    python merge_reports.py <入力フォルダ> <出力ファイル名> [--config <設定ファイル>]

例:
    python merge_reports.py ./sample_data ./output.xlsx
    python merge_reports.py ./sample_data ./output.xlsx --config my_config.json

ファイル名規則:
    各プラットフォームのCSVは、設定ファイルで指定した filename_prefix で
    始まるファイル名にしてください。デフォルトでは:
        rakuten_*.csv  → 楽天市場形式
        amazon_*.csv   → Amazon形式（商品名は日本語に変換）
        base_*.csv     → 自社BASE形式

設定ファイル（config.json）:
    各プラットフォームの列名マッピング・商品名翻訳辞書を
    JSONで自由にカスタマイズ可能。新しいプラットフォームを追加するときも、
    config.json に1ブロック追加するだけで対応可能。
"""

import argparse
import colorsys
import json
import math
import sys
import tempfile
from pathlib import Path

# Windows のターミナル（CP932）でも日本語を文字化けせず表示するためにUTF-8に切り替え
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import matplotlib

matplotlib.use("Agg")  # GUIなしのバックエンド（サーバー実行・スクリプト実行用）

import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill

# 日本語フォント設定
matplotlib.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "sans-serif"]
matplotlib.rcParams["axes.unicode_minus"] = False

# Webアプリと統一した配色（中間の彩度：くすみと派手の中間）
ACCENT_COLOR = "#4F62C4"
PLATFORM_COLORS_MPL = {
    "楽天市場": "#CC4663",          # 彩度UPローズ
    "Amazon": "#E08A47",            # 彩度UPオレンジ
    "自社BASE": "#5E6EC7",          # 彩度UPインディゴ
    "Yahoo!ショッピング": "#A572C0", # 彩度UPパープル
    "minne": "#D2649C",             # 彩度UPピンク
    "Creema": "#2E9F8E",            # 彩度UPティール
    "自社Shopify": "#58B45F",       # 彩度UPグリーン
}

# 未登録プラットフォーム用のフォールバックパレット（7色）
# 登録されていない名前にはこの中から順番に色が割り当てられる
FALLBACK_PALETTE = [
    "#CC4663",  # ローズ
    "#E08A47",  # オレンジ
    "#5E6EC7",  # インディゴ
    "#A572C0",  # パープル
    "#D2649C",  # ピンク
    "#2E9F8E",  # ティール
    "#58B45F",  # グリーン
]


def get_platform_color(
    label: str,
    position: int = 0,
    platform_config: dict | None = None,
) -> str:
    """プラットフォームの色を決定する。

    優先順位:
      1. platform_config["color"] が指定されていればその色（ユーザーカスタム）
      2. PLATFORM_COLORS_MPL に登録があればその色（既定の推奨色）
      3. FALLBACK_PALETTE から position に応じて選ぶ（未登録の補助色）
    """
    if platform_config and "color" in platform_config and platform_config["color"]:
        return platform_config["color"]
    if label in PLATFORM_COLORS_MPL:
        return PLATFORM_COLORS_MPL[label]
    return FALLBACK_PALETTE[position % len(FALLBACK_PALETTE)]


def build_platform_color_map(config: dict, platform_labels_in_order: list) -> dict:
    """プラットフォームのラベル → 色 の辞書を構築する（順位反映）。"""
    color_map: dict = {}
    # config の label → platform_config の辞書を作成
    label_to_config = {
        pconf["label"]: pconf for pconf in config.get("platforms", {}).values()
    }
    for i, label in enumerate(platform_labels_in_order):
        plat_conf = label_to_config.get(label)
        color_map[label] = get_platform_color(label, position=i, platform_config=plat_conf)
    return color_map


CANONICAL_COLUMNS = ["注文番号", "注文日", "商品名", "個数", "金額", "プラットフォーム"]


# ============================================================
# 列名の類義語辞書（自動推測用）
# ============================================================
COLUMN_SYNONYMS = {
    "注文番号": [
        "注文番号", "注文ID", "注文no", "注文NO", "注文No",
        "受注番号", "受注ID", "受注No", "受注no", "受注NO",
        "オーダーID", "オーダー番号", "オーダーNo",
        "order-id", "order_id", "order_no", "order number",
        "orderno", "orderid", "no.",
    ],
    "注文日": [
        "注文日", "注文日時", "受注日", "受注日時",
        "受付日", "受付日時", "発注日",
        "purchase-date", "order_date", "order date",
        "date", "purchase_date", "ordered_at",
    ],
    "商品名": [
        "商品名", "商品", "品名", "アイテム名", "作品名",
        "商品タイトル", "商品タイトル名",
        "product", "product-name", "product_name", "product name",
        "item", "item_name", "item-name", "item name", "title",
    ],
    "個数": [
        "個数", "数量", "数", "個",
        "qty", "quantity", "count",
    ],
    "金額": [
        "金額", "合計金額", "売上金額", "売上", "総額", "代金",
        "販売価格", "価格", "請求金額",
        "price", "total", "item-total", "total_price",
        "total price", "subtotal", "line_total", "line-total",
    ],
}


def _match_score(header: str, synonym: str) -> int:
    """ヘッダー名と類義語のマッチ度を 0-100 で返す。"""
    h = header.strip().lower()
    s = synonym.strip().lower()
    if h == s:
        return 100
    # 記号・スペースを除去した正規化マッチ
    h_norm = h.replace("-", "").replace("_", "").replace(" ", "")
    s_norm = s.replace("-", "").replace("_", "").replace(" ", "")
    if h_norm == s_norm:
        return 95
    # 部分一致（どちらかが含まれる）
    if h in s or s in h:
        return 50
    return 0


def suggest_column_mapping(headers: list[str]) -> dict:
    """CSVヘッダー一覧から、各統一列に対応するヘッダー名を推測する。

    戻り値: {統一列名: 推測したCSV列名 または None}
    貪欲法でスコア降順に割り当て、同じヘッダーが複数の統一列に使われないようにする。
    """
    scores: dict = {}
    for canonical, synonyms in COLUMN_SYNONYMS.items():
        for header in headers:
            best = 0
            for syn in synonyms:
                best = max(best, _match_score(header, syn))
            if best > 0:
                scores[(canonical, header)] = best

    sorted_pairs = sorted(scores.items(), key=lambda x: -x[1])
    used_headers: set = set()
    used_canonicals: set = set()
    suggestions: dict = {c: None for c in COLUMN_SYNONYMS.keys()}

    for (canonical, header), score in sorted_pairs:
        if canonical in used_canonicals or header in used_headers:
            continue
        if score < 50:
            continue
        suggestions[canonical] = header
        used_canonicals.add(canonical)
        used_headers.add(header)

    return suggestions


def load_config(config_path: Path) -> dict:
    """設定ファイル（JSON）を読み込んでバリデーション。"""
    if not config_path.exists():
        raise FileNotFoundError(
            f"設定ファイルが見つかりません: {config_path}\n"
            "プロジェクトルートに config.json を配置してください。"
        )
    with open(config_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    if "platforms" not in config:
        raise ValueError("設定ファイルに 'platforms' キーがありません。")

    for key, platform in config["platforms"].items():
        if "filename_prefix" not in platform:
            raise ValueError(f"プラットフォーム '{key}' に 'filename_prefix' がありません。")
        if "label" not in platform:
            raise ValueError(f"プラットフォーム '{key}' に 'label' がありません。")
        if "columns" not in platform:
            raise ValueError(f"プラットフォーム '{key}' に 'columns' マッピングがありません。")

    return config


def detect_platform(filename: str, config: dict) -> str | None:
    """ファイル名先頭からプラットフォームを判定する。"""
    lower = filename.lower()
    for platform_key, platform_config in config["platforms"].items():
        if lower.startswith(platform_config["filename_prefix"].lower()):
            return platform_key
    return None


def load_platform_csv(file: Path, platform_key: str, config: dict) -> pd.DataFrame:
    """プラットフォーム別のCSVを読み込み、統一スキーマに変換する。"""
    platform_config = config["platforms"][platform_key]
    df = pd.read_csv(file)

    # 列名を統一スキーマにリネーム
    column_map = platform_config["columns"]
    df = df.rename(columns=column_map)

    # 統一スキーマに必要な列だけ残す（余分な列はカット）
    keep_columns = [c for c in column_map.values() if c in df.columns]
    missing = [c for c in column_map.values() if c not in df.columns]
    if missing:
        print(f"  ⚠ 不足列（無視されます）: {missing}")
    df = df[keep_columns].copy()

    # 商品名の翻訳テーブルがあれば適用（例: Amazonの英語 → 日本語）
    if "product_name_translation" in platform_config and "商品名" in df.columns:
        translation = platform_config["product_name_translation"]
        df["商品名"] = df["商品名"].map(translation).fillna(df["商品名"])

    # 注文日をdatetime型に統一（文字列でもパースして揃える）
    if "注文日" in df.columns:
        df["注文日"] = pd.to_datetime(df["注文日"]).dt.strftime("%Y-%m-%d")

    # プラットフォーム列を付与
    df["プラットフォーム"] = platform_config["label"]

    # 統一スキーマの列順に並べ替え（無い列は除外）
    available = [c for c in CANONICAL_COLUMNS if c in df.columns]
    return df[available]


def load_all(folder: Path, config: dict) -> pd.DataFrame:
    """フォルダ内の対象CSVをすべて統一スキーマで読み込み、統合する。"""
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"フォルダが見つかりません: {folder}")

    frames = []
    for file in sorted(folder.iterdir()):
        if file.suffix.lower() != ".csv":
            continue

        platform_key = detect_platform(file.name, config)
        if platform_key is None:
            print(f"  ⚠ スキップ（不明な形式）: {file.name}")
            continue

        label = config["platforms"][platform_key]["label"]
        print(f"  読み込み: {file.name} → {label} として処理")
        df = load_platform_csv(file, platform_key, config)
        frames.append(df)

    if not frames:
        prefixes = [p["filename_prefix"] for p in config["platforms"].values()]
        raise ValueError(
            "対象CSVが見つかりませんでした。\n"
            f"ファイル名は次のいずれかで始まる必要があります: {prefixes}"
        )

    return pd.concat(frames, ignore_index=True)


def remove_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    """注文番号ベースで重複削除（万一の二重エクスポート対策）。"""
    before = len(df)
    df = df.drop_duplicates(subset=["注文番号"]).reset_index(drop=True)
    after = len(df)
    if before != after:
        print(f"  重複削除: {before} 行 → {after} 行 ({before - after} 行削除)")
    return df


def build_product_summary(df: pd.DataFrame) -> pd.DataFrame:
    """商品別の売上集計を作る。"""
    summary = (
        df.groupby("商品名", as_index=False)
        .agg(注文件数=("注文番号", "count"), 販売個数=("個数", "sum"), 売上合計=("金額", "sum"))
        .sort_values("売上合計", ascending=False)
        .reset_index(drop=True)
    )
    return summary


def build_platform_summary(df: pd.DataFrame) -> pd.DataFrame:
    """プラットフォーム別の売上集計を作る。"""
    summary = (
        df.groupby("プラットフォーム", as_index=False)
        .agg(注文件数=("注文番号", "count"), 販売個数=("個数", "sum"), 売上合計=("金額", "sum"))
        .sort_values("売上合計", ascending=False)
        .reset_index(drop=True)
    )
    total = summary["売上合計"].sum()
    summary["売上構成比"] = (summary["売上合計"] / total * 100).round(1).astype(str) + "%"
    return summary


def build_daily_summary(df: pd.DataFrame) -> pd.DataFrame:
    """日別の売上推移を作る。"""
    summary = (
        df.groupby("注文日", as_index=False)
        .agg(注文件数=("注文番号", "count"), 売上合計=("金額", "sum"))
        .sort_values("注文日")
        .reset_index(drop=True)
    )
    return summary


def style_header(ws, color: str = "4F46E5") -> None:
    """ヘッダー行に色付け・白太字、ヘッダー行を固定。"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26  # ヘッダー行を少し高く
    ws.freeze_panes = "A2"


def _visual_width(value) -> float:
    """文字列の見た目の幅を見積もる（日本語は約2倍幅、英数字は1倍）。"""
    if value is None:
        return 0
    width = 0.0
    for ch in str(value):
        if ord(ch) > 127:
            width += 2.2
        else:
            width += 1.15
    return width


def autosize_columns(ws) -> None:
    """列幅を内容に合わせて自動調整する（日本語の表示幅を考慮）。"""
    for column in ws.columns:
        max_w = max(
            (_visual_width(cell.value) for cell in column if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[column[0].column_letter].width = min(max_w + 3, 55)


def format_number_columns(ws, currency_cols: list[int], qty_cols: list[int], max_row: int) -> None:
    """通貨列(¥#,##0) と 数量列(#,##0) のフォーマットを適用する。"""
    for col_idx in currency_cols:
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"¥"#,##0'
    for col_idx in qty_cols:
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def apply_alternating_rows(ws, max_row: int, max_col: int, color: str = "F9FAFB") -> None:
    """偶数行に薄い背景色を入れて読みやすさUP。"""
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for row_idx in range(2, max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill


# ============================================================
# matplotlib によるグラフ画像生成（Web アプリと統一感のあるデザイン）
# ============================================================
def _adjust_color_by_rank(hex_color: str, rank: int, total: int) -> str:
    """ランクが大きい（=小さなスライス）ほど色を控えめにする。

    rank=0 → 元の色そのまま（売上 No.1 の最も主張する色）
    rank=total-1 → 彩度を 55% 減・明度を 30% 白寄りに（最も控えめ）
    """
    hex_color = hex_color.lstrip("#")
    r = int(hex_color[0:2], 16) / 255
    g = int(hex_color[2:4], 16) / 255
    b = int(hex_color[4:6], 16) / 255
    h, l, s = colorsys.rgb_to_hls(r, g, b)

    factor = rank / (total - 1) if total > 1 else 0
    s_new = max(0.0, s * (1 - factor * 0.55))
    l_new = l + (1 - l) * factor * 0.30

    r2, g2, b2 = colorsys.hls_to_rgb(h, l_new, s_new)
    return f"#{int(r2*255):02X}{int(g2*255):02X}{int(b2*255):02X}"



def _render_product_chart(product_df: pd.DataFrame, output_path: Path) -> None:
    """商品別売上の横棒グラフ（グラデーション付き）。"""
    df = product_df.copy().sort_values("売上合計", ascending=True)
    n = len(df)

    fig, ax = plt.subplots(figsize=(11, max(4.5, 0.55 * n + 1.5)))

    # 売上に応じた紫グラデーションを生成
    cmap = mcolors.LinearSegmentedColormap.from_list("indigo_grad", ["#C7D2FE", "#4F46E5"])
    if df["売上合計"].max() == df["売上合計"].min():
        colors = [ACCENT_COLOR] * n
    else:
        norm = plt.Normalize(df["売上合計"].min(), df["売上合計"].max())
        colors = [cmap(norm(v)) for v in df["売上合計"]]

    bars = ax.barh(df["商品名"], df["売上合計"], color=colors, edgecolor="none")

    max_val = df["売上合計"].max()
    for bar, val in zip(bars, df["売上合計"]):
        ax.text(
            val + max_val * 0.012,
            bar.get_y() + bar.get_height() / 2,
            f"¥{int(val):,}",
            va="center",
            fontsize=10,
            color="#1F2937",
            fontweight="bold",
        )

    ax.set_xlabel("売上合計（円）", fontsize=11, color="#4B5563")
    ax.set_title("商品別 売上ランキング", fontsize=15, fontweight="bold", pad=18, color="#111827")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#E5E7EB")
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.tick_params(colors="#6B7280")
    ax.set_axisbelow(True)
    ax.grid(axis="x", alpha=0.35, color="#E5E7EB")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, p: f"¥{int(v):,}"))
    ax.set_xlim(0, max_val * 1.18)
    plt.tight_layout()
    plt.savefig(output_path, dpi=130, bbox_inches="tight", facecolor="white")
    plt.close()


def _render_platform_donut(
    platform_df: pd.DataFrame,
    output_path: Path,
    color_map: dict | None = None,
) -> None:
    """プラットフォーム別売上構成のドーナツチャート（引き出し線付きラベル + 中央合計）。

    color_map: ラベル → ベース色 の辞書。指定が無ければラベルから推測。
    """
    fig, ax = plt.subplots(figsize=(8.5, 7))

    n_platforms = len(platform_df)
    # color_map から各プラットフォームのベース色を取得し、ランクに応じて薄め調整
    colors = []
    for i, p_label in enumerate(platform_df["プラットフォーム"]):
        if color_map and p_label in color_map:
            base = color_map[p_label]
        else:
            base = get_platform_color(p_label, position=i)
        colors.append(_adjust_color_by_rank(base, rank=i, total=n_platforms))
    total = platform_df["売上合計"].sum()

    wedges, _ = ax.pie(
        platform_df["売上合計"],
        labels=None,
        colors=colors,
        startangle=90,
        counterclock=False,
        wedgeprops={"edgecolor": "white", "linewidth": 2, "width": 0.4},
    )

    # 引き出し線とラベル（固定オフセットで全スライス同じ曲がり角度に）
    for i, wedge in enumerate(wedges):
        mid_angle_deg = (wedge.theta1 + wedge.theta2) / 2
        ang = math.radians(mid_angle_deg)
        cos_a = math.cos(ang)
        sin_a = math.sin(ang)

        # ① パイ外周
        x0, y0 = cos_a, sin_a

        # ② ヒジ：固定オフセットで曲がりを保証（水平に近いスライスでも見える）
        horizontal_dir = 1 if cos_a >= 0 else -1
        vertical_dir = 1 if sin_a >= 0 else -1
        bx = x0 + 0.13 * horizontal_dir
        by = y0 + 0.22 * vertical_dir

        # ③ ラベル位置（ヒジから水平に延長）
        x_offset = 0.22 if cos_a >= 0 else -0.22
        lx = bx + x_offset
        ly = by

        # 引き出し線（2セグメント）
        ax.plot([x0, bx, lx], [y0, by, ly], color="#6B7280", linewidth=1)

        # テキスト
        ha = "left" if cos_a >= 0 else "right"
        p_name = platform_df["プラットフォーム"].iloc[i]
        percent = platform_df["売上合計"].iloc[i] / total * 100
        slice_color = colors[i]
        text_x = lx + (0.03 if cos_a >= 0 else -0.03)
        ax.text(
            text_x, ly + 0.07,
            p_name,
            fontsize=11, color="#374151", ha=ha, va="center", fontweight="bold",
        )
        ax.text(
            text_x, ly - 0.09,
            f"{percent:.1f}%",
            fontsize=13, color=slice_color, ha=ha, va="center", fontweight="bold",
        )

    # 中央テキスト
    total_orders = int(platform_df["注文件数"].sum())
    ax.text(0, 0.18, "売上合計", ha="center", va="center", fontsize=10, color="#4B5563",
            fontweight="bold")
    ax.text(0, 0, f"¥{int(total):,}", ha="center", va="center",
            fontsize=17, fontweight="bold", color="#111827")
    ax.text(0, -0.18, f"{total_orders}件", ha="center", va="center",
            fontsize=10, color="#4B5563")

    ax.set_title("プラットフォーム別 売上構成", fontsize=15, fontweight="bold", pad=20, color="#111827")
    ax.set_xlim(-2.0, 2.0)
    ax.set_ylim(-1.6, 1.6)
    ax.set_aspect("equal")
    ax.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=130, bbox_inches="tight", facecolor="white")
    plt.close()


def _render_daily_trend(daily_df: pd.DataFrame, output_path: Path) -> None:
    """日別売上推移の折れ線+塗りつぶしチャート。"""
    fig, ax = plt.subplots(figsize=(12, 5))

    x = list(range(len(daily_df)))
    y = daily_df["売上合計"].values

    ax.fill_between(x, y, color=ACCENT_COLOR, alpha=0.13)
    ax.plot(x, y, marker="o", linewidth=2.5, color=ACCENT_COLOR,
            markersize=7, markerfacecolor="white", markeredgewidth=2)

    ax.set_xticks(x)
    ax.set_xticklabels(daily_df["注文日"].astype(str), rotation=45, ha="right")
    ax.set_xlabel("注文日", fontsize=11, color="#4B5563")
    ax.set_ylabel("売上合計（円）", fontsize=11, color="#4B5563")
    ax.set_title("日別 売上推移", fontsize=15, fontweight="bold", pad=18, color="#111827")

    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, p: f"¥{int(v):,}"))

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color("#E5E7EB")
    ax.spines["bottom"].set_color("#E5E7EB")
    ax.tick_params(colors="#6B7280")
    ax.set_axisbelow(True)
    ax.grid(alpha=0.35, color="#E5E7EB")
    ax.margins(x=0.02)

    plt.tight_layout()
    plt.savefig(output_path, dpi=130, bbox_inches="tight", facecolor="white")
    plt.close()


def write_excel(
    df: pd.DataFrame,
    product_summary: pd.DataFrame,
    platform_summary: pd.DataFrame,
    daily_summary: pd.DataFrame,
    output: Path,
    config: dict | None = None,
) -> None:
    """4シート構成のExcelを出力する（フォーマット・グラフ強化版）。

    config を渡すと、各プラットフォームの "color" 設定が円グラフに反映される。
    """
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="全注文", index=False)
        product_summary.to_excel(writer, sheet_name="商品別集計", index=False)
        platform_summary.to_excel(writer, sheet_name="プラットフォーム別集計", index=False)
        daily_summary.to_excel(writer, sheet_name="日別売上推移", index=False)

    wb = load_workbook(output)

    # =============================
    # 全注文シート
    # =============================
    ws_all = wb["全注文"]
    style_header(ws_all)
    autosize_columns(ws_all)
    df_cols = list(df.columns)
    currency_cols = [df_cols.index("金額") + 1] if "金額" in df_cols else []
    qty_cols = [df_cols.index("個数") + 1] if "個数" in df_cols else []
    format_number_columns(ws_all, currency_cols, qty_cols, max_row=len(df) + 1)
    apply_alternating_rows(ws_all, max_row=len(df) + 1, max_col=len(df.columns))

    # =============================
    # 商品別集計シート + 横棒グラフ
    # =============================
    ws_product = wb["商品別集計"]
    style_header(ws_product)
    autosize_columns(ws_product)
    format_number_columns(
        ws_product, currency_cols=[4], qty_cols=[2, 3], max_row=len(product_summary) + 1
    )
    apply_alternating_rows(ws_product, max_row=len(product_summary) + 1, max_col=4)

    # =============================
    # プラットフォーム別集計シート（スタイリングのみ、グラフは下で埋め込む）
    # =============================
    ws_plat = wb["プラットフォーム別集計"]
    style_header(ws_plat)
    autosize_columns(ws_plat)
    format_number_columns(
        ws_plat, currency_cols=[4], qty_cols=[2, 3], max_row=len(platform_summary) + 1
    )
    apply_alternating_rows(ws_plat, max_row=len(platform_summary) + 1, max_col=5)

    # =============================
    # 日別売上推移シート（スタイリングのみ、グラフは下で埋め込む）
    # =============================
    ws_daily = wb["日別売上推移"]
    style_header(ws_daily)
    autosize_columns(ws_daily)
    format_number_columns(
        ws_daily, currency_cols=[3], qty_cols=[2], max_row=len(daily_summary) + 1
    )
    apply_alternating_rows(ws_daily, max_row=len(daily_summary) + 1, max_col=3)

    # =============================
    # グラフを matplotlib で生成して PNG として埋め込み（Webアプリと統一感のあるデザイン）
    # =============================
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)

        if len(product_summary) > 0:
            product_png = tmp_path / "product_chart.png"
            _render_product_chart(product_summary, product_png)
            img = XLImage(str(product_png))
            img.width, img.height = 740, max(360, 50 + 36 * len(product_summary))
            ws_product.add_image(img, "F2")

        if len(platform_summary) > 0:
            platform_png = tmp_path / "platform_chart.png"
            # configからユーザー設定の色を取得（無ければ既定色を使用）
            color_map = (
                build_platform_color_map(
                    config, list(platform_summary["プラットフォーム"])
                )
                if config
                else None
            )
            _render_platform_donut(platform_summary, platform_png, color_map=color_map)
            img = XLImage(str(platform_png))
            img.width, img.height = 620, 520
            ws_plat.add_image(img, "G2")

        if len(daily_summary) > 0:
            daily_png = tmp_path / "daily_chart.png"
            _render_daily_trend(daily_summary, daily_png)
            img = XLImage(str(daily_png))
            img.width, img.height = 880, 380
            ws_daily.add_image(img, "E2")

        wb.save(output)


def print_summary_to_console(
    df: pd.DataFrame,
    product_summary: pd.DataFrame,
    platform_summary: pd.DataFrame,
) -> None:
    """ターミナルにも結果のサマリーを表示する。"""
    total_orders = len(df)
    total_units = int(df["個数"].sum())
    total_revenue = int(df["金額"].sum())

    print()
    print("=" * 50)
    print("[集計サマリー]")
    print("=" * 50)
    print(f"  総注文数:   {total_orders} 件")
    print(f"  総販売個数: {total_units} 個")
    print(f"  売上合計:   ¥{total_revenue:,}")
    print()
    print("  プラットフォーム別:")
    for _, row in platform_summary.iterrows():
        print(
            f"    {row['プラットフォーム']:<10} "
            f"{row['注文件数']:>3}件 / ¥{int(row['売上合計']):>8,} ({row['売上構成比']})"
        )
    print()
    print("  売上TOP3:")
    for i, (_, row) in enumerate(product_summary.head(3).iterrows(), start=1):
        print(f"    {i}. {row['商品名']:<30} ¥{int(row['売上合計']):>7,}")
    print("=" * 50)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="EC事業者向け 受注データ統合・集計ツール",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input_folder", help="入力CSVが入っているフォルダ")
    parser.add_argument("output_file", help="出力するExcelファイル名")
    parser.add_argument(
        "--config",
        default="config.json",
        help="設定ファイルのパス（既定: config.json）",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    input_folder = Path(args.input_folder)
    output_file = Path(args.output_file)
    config_path = Path(args.config)

    print(f"[0/5] 設定ファイル読み込み: {config_path}")
    config = load_config(config_path)
    print(f"  対応プラットフォーム: {list(config['platforms'].keys())}")

    print(f"[1/5] ファイル読み込み: {input_folder}")
    df = load_all(input_folder, config)
    print(f"  → 合計 {len(df)} 行")

    print("[2/5] データクリーニング")
    df = remove_duplicates(df)

    print("[3/5] 集計シート生成")
    product_summary = build_product_summary(df)
    platform_summary = build_platform_summary(df)
    daily_summary = build_daily_summary(df)

    print(f"[4/5] Excel出力: {output_file}")
    write_excel(df, product_summary, platform_summary, daily_summary, output_file, config=config)

    print("[5/5] 完了！")
    print_summary_to_console(df, product_summary, platform_summary)
    print(f"\n  Excelファイル: {output_file.resolve()}")


if __name__ == "__main__":
    main()
